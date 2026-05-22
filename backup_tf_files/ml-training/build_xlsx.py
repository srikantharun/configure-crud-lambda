"""
build_xlsx.py
=============
Convert validate.py JSON evidence files into the 17-column XLSX schema that
policy_scorer.py expects. One XLSX is produced per (policy, placement).

Defaults:
  Input  : runs/<RUN_DATE>_run/validate_baseline_<policy>_<placement>.json
  Output : runs/validate_baseline_<policy>_<placement>_<RUN_DATE>.xlsx

Each XLSX contains three sheets:
  1. main sheet (17 columns) - the spreadsheet your manager sees
  2. Audit tab - request_id, CW Insights query, log groups, terminating rule
  3. Sanity check tab - summary counts, traceability notes

Override defaults at the command line:
  python build_xlsx.py                                    # uses RUN_DATE default
  python build_xlsx.py --date 20260520                    # explicit date
  python build_xlsx.py --src-dir /path/to/json --date 20260520
  python build_xlsx.py --out-dir runs --date 20260520
"""
import argparse
import json
import os
import re
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

DEFAULT_RUN_DATE = "20260520"

LOOKUP_SRC_CANDIDATES = [
    os.path.join(HERE, "runs", "validate_baseline_1.3a_post_20260517.xlsx"),
    os.path.join(HERE, "runs", "validate_baseline_1.1_post_20260518.xlsx"),
]

HEADERS_MAIN = [
    "NSEC / OWASP Category",
    "Test ID / JuiceShop ID",
    "Result 1.3",
    "Attack Payload",
    "Baseline 1.3",
    "Legacy Baseline",
    "1.0 Baseline",
    "1.1 Baseline",
    "Baseline 1.3 Evidence where Blocked",
    "Security Misconfiguration Observed - Was a Security Gap Introduced where blocked previously?",
    "Where Security Misconfiguration",
    "Where False Negative not due to Misconfiguration",
    "Where False Negative Successful Exploit",
    "Where False Negative Research links (OWASP/NIST/MITRE etc.)",
    "Where False Negative Risk Probability of Exploit",
    "Where False Negative Risk Impact of Exploit",
    "Request ID",
]

HEADERS_AUDIT = [
    "Test ID", "Request ID", "Verdict",
    "CF terminating rule", "CF rule group", "ALB terminating rule",
    "Payload", "CW Insights query (CF)", "CW Insights query (ALB)",
    "CF log group", "ALB log group", "CF action", "CF timestamp",
]

POLICIES = ["1.1", "1.3", "1.3b"]
PLACEMENTS = [
    ("POST",       "post"),
    ("GET-QS",     "get_qs"),
    ("GET-Cookie", "get_cookie"),
    ("GET-URI",    "get_uri"),
    ("GET-Header", "get_header"),
]


def load_lookup():
    """Read test_id -> (category, payload) from a known-good xlsx."""
    for path in LOOKUP_SRC_CANDIDATES:
        if os.path.exists(path):
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            out = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                tid = row[1]
                if tid is None:
                    continue
                out[str(tid)] = {"category": row[0] or "", "payload": row[3] or ""}
            wb.close()
            return out
    print("WARN: no lookup xlsx found, category/payload columns will be blank")
    return {}


def parse_rule(terminating_rule_id):
    if not terminating_rule_id:
        return "", ""
    m = re.match(r"PREFMManaged-(.+?)-\d+$", terminating_rule_id)
    if m:
        return terminating_rule_id, m.group(1).replace("AWSManagedRules", "")
    return terminating_rule_id, terminating_rule_id


def cw_insights_query(request_id):
    if not request_id:
        return ""
    return (f"fields @timestamp, action, terminatingRuleId, terminatingRuleType "
            f"| filter @message like /{request_id}/ "
            f"| sort @timestamp desc | limit 5")


def verdict_color(verdict):
    v = (verdict or "").upper()
    if v == "CF_BLOCK":             return PatternFill("solid", fgColor="C6EFCE")
    if v == "CF_ALLOW_ALB_BLOCK":   return PatternFill("solid", fgColor="FFEB9C")
    if v == "CF_ALLOW_ALB_ALLOW":   return PatternFill("solid", fgColor="FFC7CE")
    if v == "NO_CF_LOG":            return PatternFill("solid", fgColor="D9D9D9")
    return None


def build(json_path, out_path, policy_label, variant_label, lookup, run_date):
    raw = open(json_path).read()
    data = json.loads(raw, strict=False)
    evidence = data.get("evidence", [])
    cf_log_group = data.get("cf_log_group", "")
    alb_log_group = data.get("alb_log_group", "")

    wb = Workbook()
    sheet_name = f"Baseline {policy_label} {variant_label}"[:31]
    ws = wb.active
    ws.title = sheet_name

    for c, h in enumerate(HEADERS_MAIN, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = PatternFill("solid", fgColor="305496")

    def sort_key(rec):
        rid = rec.get("requirement_id", "")
        m = re.search(r"juiceshop-(\d+)", str(rid))
        return int(m.group(1)) if m else 99999
    evidence_sorted = sorted(evidence, key=sort_key)

    for r_idx, rec in enumerate(evidence_sorted, start=2):
        rid = str(rec.get("requirement_id") or "")
        lk = lookup.get(rid, {})
        cf = rec.get("cf_waf_log") or {}
        verdict = rec.get("verdict", "")
        rule_id, _ = parse_rule(cf.get("terminating_rule_id"))
        ev_text = rec.get("evidence", "")

        ws.cell(row=r_idx, column=1, value=lk.get("category", ""))
        ws.cell(row=r_idx, column=2, value=rid)
        c3 = ws.cell(row=r_idx, column=3, value=verdict)
        fill = verdict_color(verdict)
        if fill: c3.fill = fill
        ws.cell(row=r_idx, column=4, value=lk.get("payload", ""))
        ws.cell(row=r_idx, column=5, value=ev_text)
        ws.cell(row=r_idx, column=9, value=rule_id)
        ws.cell(row=r_idx, column=17, value=rec.get("request_id", ""))

    widths = [18, 22, 16, 50, 60, 10, 10, 10, 50, 22, 22, 22, 22, 22, 18, 18, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"

    a = wb.create_sheet("Audit")
    for c, h in enumerate(HEADERS_AUDIT, 1):
        cell = a.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for r_idx, rec in enumerate(evidence_sorted, start=2):
        rid = str(rec.get("requirement_id") or "")
        cf = rec.get("cf_waf_log") or {}
        alb = rec.get("alb_waf_log") or {}
        verdict = rec.get("verdict", "")
        rule_id, rule_short = parse_rule(cf.get("terminating_rule_id"))
        alb_rule_id, _ = parse_rule(alb.get("terminating_rule_id"))
        req_id = rec.get("request_id", "")
        lk = lookup.get(rid, {})
        a.cell(row=r_idx, column=1, value=rid)
        a.cell(row=r_idx, column=2, value=req_id)
        c3 = a.cell(row=r_idx, column=3, value=verdict)
        fill = verdict_color(verdict)
        if fill: c3.fill = fill
        a.cell(row=r_idx, column=4, value=rule_id)
        a.cell(row=r_idx, column=5, value=rule_short)
        a.cell(row=r_idx, column=6, value=alb_rule_id)
        a.cell(row=r_idx, column=7, value=lk.get("payload", ""))
        a.cell(row=r_idx, column=8, value=cw_insights_query(req_id))
        a.cell(row=r_idx, column=9, value=cw_insights_query(req_id))
        a.cell(row=r_idx, column=10, value=cf_log_group)
        a.cell(row=r_idx, column=11, value=alb_log_group)
        a.cell(row=r_idx, column=12, value=cf.get("action", ""))
        a.cell(row=r_idx, column=13, value=cf.get("timestamp", ""))
    widths_a = [20, 45, 22, 60, 30, 50, 50, 80, 80, 25, 25, 12, 22]
    for i, w in enumerate(widths_a, 1):
        a.column_dimensions[get_column_letter(i)].width = w
    a.freeze_panes = "C2"

    s = wb.create_sheet("Sanity check")
    summary = data.get("summary", {})
    rows = [
        ("Source JSON", os.path.basename(json_path)),
        ("Policy label", policy_label),
        ("Variant", variant_label),
        ("Generated at", f"{run_date[:4]}-{run_date[4:6]}-{run_date[6:]}"),
        ("Total records", summary.get("total_validated", "")),
        ("CF_BLOCK", summary.get("cf_block", "")),
        ("CF_ALLOW_ALB_BLOCK", summary.get("cf_allow_alb_block", "")),
        ("CF_ALLOW_ALB_ALLOW", summary.get("cf_allow_alb_allow", "")),
        ("NO_CF_LOG", summary.get("no_cf_log", "")),
        ("", ""),
        ("Verification:", "Compare totals here to row count on main sheet"),
        ("Audit tab:", "Use CW Insights query column to verify any test_id in CloudWatch"),
    ]
    for r_idx, (k, v) in enumerate(rows, start=1):
        c = s.cell(row=r_idx, column=1, value=k); c.font = Font(bold=True)
        s.cell(row=r_idx, column=2, value=v)
    s.column_dimensions["A"].width = 22
    s.column_dimensions["B"].width = 60

    wb.save(out_path)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--date", default=DEFAULT_RUN_DATE,
                   help="run date YYYYMMDD (default %(default)s)")
    p.add_argument("--src-dir", default=None,
                   help="json source dir (default: runs/<date>_run/)")
    p.add_argument("--out-dir", default=os.path.join(HERE, "runs"),
                   help="output dir for xlsx (default: runs/)")
    args = p.parse_args()

    src_dir = args.src_dir or os.path.join(HERE, "runs", f"{args.date}_run")
    if not os.path.isdir(src_dir):
        print(f"ERROR: source dir {src_dir} not found")
        sys.exit(1)
    os.makedirs(args.out_dir, exist_ok=True)

    lookup = load_lookup()
    print(f"Lookup loaded: {len(lookup)} test_ids")
    print(f"Source: {src_dir}")
    print(f"Output: {args.out_dir}")
    print()

    n_ok, n_err, n_skip = 0, 0, 0
    for pol in POLICIES:
        for var_label, var_slug in PLACEMENTS:
            src = os.path.join(src_dir, f"validate_baseline_{pol}_{var_slug}.json")
            if not os.path.exists(src):
                print(f"  skip {pol}_{var_slug}  (missing {os.path.basename(src)})")
                n_skip += 1
                continue
            out = os.path.join(args.out_dir, f"validate_baseline_{pol}_{var_slug}_{args.date}.xlsx")
            try:
                build(src, out, pol, var_label, lookup, args.date)
                print(f"  OK   {os.path.basename(out)}")
                n_ok += 1
            except Exception as e:
                print(f"  ERR  {os.path.basename(src)} -> {e}")
                n_err += 1
    print()
    print(f"Built {n_ok} XLSXs ({n_err} errors, {n_skip} skipped)")


if __name__ == "__main__":
    main()
