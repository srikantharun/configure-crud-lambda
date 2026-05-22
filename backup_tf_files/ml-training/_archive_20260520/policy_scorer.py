"""
policy_scorer.py
================
Reads `runs/validate_<policy>_<variant>_<YYYYMMDD>.xlsx` files (17-col schema,
verdict in column C = "Result 1.3"), aggregates per (policy, variant) and per
policy, detects artefact-duplicate files (identical verdict maps across
different placements), and writes CSV scorecards.

Outputs (in output/):
  per_run_scorecard.csv         one row per xlsx file
  policy_variant_scorecard.csv  mean/std per (policy, variant)
  policy_scorecard.csv          per-policy aggregate (input to mcda_decider.py)
  artefacts.txt                 duplicate-fingerprint report
"""
import csv
import glob
import hashlib
import os
import re
import statistics
from collections import Counter, defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
RUNS_DIR = os.path.join(HERE, "runs")
OUT_DIR = os.path.join(HERE, "output")

FNAME_RE = re.compile(
    r"validate_(?P<policy>baseline_[\d\.]+[a-z]?)"
    r"_(?P<variant>post|get_qs|get_cookie|get_uri|get_header)"
    r"_(?P<date>\d{8})\.xlsx$"
)

VERDICT_COL_IDX = 2   # column C "Result 1.3"
TESTID_COL_IDX = 1    # column B "Test ID / JuiceShop ID"
REQID_COL_IDX = 16    # column Q "Request ID"


def parse_run(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    verdicts = {}
    request_ids = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tid = row[TESTID_COL_IDX]
        v = row[VERDICT_COL_IDX]
        if tid is None or v is None:
            continue
        verdicts[str(tid)] = str(v)
        if len(row) > REQID_COL_IDX and row[REQID_COL_IDX]:
            request_ids[str(tid)] = str(row[REQID_COL_IDX])
    wb.close()
    return verdicts, request_ids


def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    runs = []
    for path in sorted(glob.glob(os.path.join(RUNS_DIR, "*.xlsx"))):
        fn = os.path.basename(path)
        m = FNAME_RE.search(fn)
        if not m:
            print(f"  skip (bad filename): {fn}")
            continue
        verdicts, request_ids = parse_run(path)
        items = tuple(sorted(verdicts.items()))
        fp = hashlib.sha256(repr(items).encode()).hexdigest()[:12]
        rid_items = tuple(sorted(request_ids.items()))
        rid_fp = hashlib.sha256(repr(rid_items).encode()).hexdigest()[:12] if rid_items else "no-rids"
        runs.append({
            "policy": m.group("policy"),
            "variant": m.group("variant"),
            "date": m.group("date"),
            "path": path,
            "verdicts": verdicts,
            "request_ids": request_ids,
            "fp": fp,
            "rid_fp": rid_fp,
        })

    # Artefact detection: SAME REQUEST IDs across different (policy,variant) tuples.
    # Request IDs are the authoritative signal of a file copy (same execution).
    # Verdict-only collisions are noted separately as "suspect coincidences".
    rid_groups = defaultdict(list)
    for r in runs:
        if r["rid_fp"] == "no-rids":
            continue
        rid_groups[r["rid_fp"]].append(r)
    artefact_groups = []
    artefact_paths = set()
    for rid_fp, grp in rid_groups.items():
        keys = {(r["policy"], r["variant"]) for r in grp}
        if len(keys) > 1:
            artefact_groups.append(grp)
            for r in grp:
                artefact_paths.add(r["path"])

    # Verdict-only coincidences (not artefacts but worth tracking)
    fp_groups = defaultdict(list)
    for r in runs:
        fp_groups[r["fp"]].append(r)
    coincidence_groups = []
    for fp, grp in fp_groups.items():
        keys = {(r["policy"], r["variant"]) for r in grp}
        if len(keys) > 1:
            # Only count as coincidence if not already an artefact
            paths_in_grp = {r["path"] for r in grp}
            if not paths_in_grp.intersection(artefact_paths):
                coincidence_groups.append(grp)

    # Per-file row
    per_run = []
    for r in runs:
        cnt = Counter(r["verdicts"].values())
        total = sum(cnt.values())
        blk = cnt.get("CF_BLOCK", 0)
        alb = cnt.get("CF_ALLOW_ALB_BLOCK", 0)
        fail = cnt.get("CF_ALLOW_ALB_ALLOW", 0)
        nocf = cnt.get("NO_CF_LOG", 0)
        per_run.append({
            "policy": r["policy"],
            "variant": r["variant"],
            "date": r["date"],
            "total": total,
            "cf_block": blk,
            "cf_allow_alb_block": alb,
            "fail": fail,
            "no_cf_log": nocf,
            "pure_block_pct": round(blk / total * 100, 2) if total else 0.0,
            "did_pct": round(alb / total * 100, 2) if total else 0.0,
            "fail_pct": round(fail / total * 100, 2) if total else 0.0,
            "artefact_flag": r["path"] in artefact_paths,
            "fingerprint": r["fp"],
            "file": os.path.basename(r["path"]),
        })

    per_run_csv = os.path.join(OUT_DIR, "per_run_scorecard.csv")
    with open(per_run_csv, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(per_run[0].keys()))
        w.writeheader()
        w.writerows(per_run)

    # Per (policy, variant) aggregate -- artefacts excluded
    pv_key = defaultdict(list)
    for r in per_run:
        if r["artefact_flag"]:
            continue
        pv_key[(r["policy"], r["variant"])].append(r["pure_block_pct"])
    pv_rows = []
    for (pol, var), vals in sorted(pv_key.items()):
        pv_rows.append({
            "policy": pol,
            "variant": var,
            "runs": len(vals),
            "mean_pure_block": round(statistics.mean(vals), 2),
            "std_pure_block": round(statistics.stdev(vals), 2) if len(vals) > 1 else 0.0,
            "min_pure_block": round(min(vals), 2),
            "max_pure_block": round(max(vals), 2),
        })
    pv_csv = os.path.join(OUT_DIR, "policy_variant_scorecard.csv")
    with open(pv_csv, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(pv_rows[0].keys()))
        w.writeheader()
        w.writerows(pv_rows)

    # Per-policy aggregate -- mean and std across variants
    pol_key = defaultdict(list)
    for r in pv_rows:
        pol_key[r["policy"]].append(r["mean_pure_block"])
    pol_rows = []
    for pol, means in sorted(pol_key.items()):
        pol_rows.append({
            "policy": pol,
            "variants_scored": len(means),
            "catch_mean_across_variants": round(statistics.mean(means), 2),
            "catch_std_across_variants": round(statistics.stdev(means), 2) if len(means) > 1 else 0.0,
            "catch_min_variant": round(min(means), 2),
            "catch_max_variant": round(max(means), 2),
        })
    pol_csv = os.path.join(OUT_DIR, "policy_scorecard.csv")
    with open(pol_csv, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(pol_rows[0].keys()))
        w.writeheader()
        w.writerows(pol_rows)

    # Artefact report
    art_txt = os.path.join(OUT_DIR, "artefacts.txt")
    with open(art_txt, "w") as fh:
        fh.write("ARTEFACT DETECTION REPORT\n")
        fh.write("=" * 60 + "\n\n")
        fh.write("PART 1 - CONFIRMED FILE-COPY ARTEFACTS (excluded from aggregates)\n")
        fh.write("Detection signal: identical Request-ID set across files.\n")
        fh.write("Request IDs come from validate.py per execution; matching IDs\n")
        fh.write("across different (policy, variant) tuples mean the same execution\n")
        fh.write("was saved twice. These files are EXCLUDED from aggregates.\n\n")
        if not artefact_groups:
            fh.write("  No file-copy artefacts detected.\n")
        else:
            for i, grp in enumerate(artefact_groups, 1):
                fh.write(f"Group {i}  (request-id-set fingerprint {grp[0]['rid_fp']})\n")
                fh.write("-" * 60 + "\n")
                for r in grp:
                    fh.write(f"  {r['policy']:<20} {r['variant']:<12} {r['date']}  ({os.path.basename(r['path'])})\n")
                fh.write("\n")
        fh.write("\nPART 2 - VERDICT-ONLY COINCIDENCES (not excluded, flagged for review)\n")
        fh.write("These files have IDENTICAL per-test-id verdict maps but DIFFERENT\n")
        fh.write("Request IDs - so they are separately-executed runs that happened to\n")
        fh.write("produce the same verdict outcomes. Not artefacts. Counted normally.\n\n")
        if not coincidence_groups:
            fh.write("  No verdict-only coincidences detected.\n")
        else:
            for i, grp in enumerate(coincidence_groups, 1):
                fh.write(f"Group {i}  (verdict-map fingerprint {grp[0]['fp']})\n")
                fh.write("-" * 60 + "\n")
                for r in grp:
                    fh.write(f"  {r['policy']:<20} {r['variant']:<12} {r['date']}  ({os.path.basename(r['path'])})\n")
                fh.write("\n")

    print(f"Per-run scorecard       : {per_run_csv}  ({len(per_run)} rows)")
    print(f"Per-policy-variant scorecard : {pv_csv}  ({len(pv_rows)} rows)")
    print(f"Per-policy scorecard    : {pol_csv}  ({len(pol_rows)} rows)")
    print(f"Artefact report         : {art_txt}  ({len(artefact_groups)} group(s))")


if __name__ == "__main__":
    main()
