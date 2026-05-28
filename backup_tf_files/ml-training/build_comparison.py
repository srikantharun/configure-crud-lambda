"""
build_comparison.py
===================
Generate the manager-facing comparison workbook that pivots
validate_baseline_*.json across 1.1 / 1.3 / 1.3b and all 6 placements
(POST-QS, POST-Body, GET-QS, GET-Cookie, GET-URI, GET-Header)
into a single XLSX with:

  - Definitions tab    how to read the workbook + known caveats
  - Summary tab        per-ruleset counts (COUNTIF) and CF_BLOCK share
                       by ruleset / by ruleset x placement
  - 11 per-ruleset tabs Common, SQLi, KnownBadInputs, Windows, Linux,
                       Unix, PHP, AdminProtection, WordPress, cyberwasp,
                       owasp-top10  each row is a test_id; cells are
                       nested-IF formulas resolved at open time
  - 18 RAW tabs        one per (policy, placement) feeding the formulas

Cells are colored green (both / 1.3 only) or red (1.1 only REGRESSION).
Attribution shifts are intentionally blank.

Usage
-----
  python build_comparison.py --date 20260521
  python build_comparison.py --date 20260521 --src-dir ~/Downloads
  python build_comparison.py --date 20260521 --out-dir ~/Downloads
"""
import argparse
import json
import os
import re
import sys
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties

HERE = os.path.dirname(os.path.abspath(__file__))

PLACEMENTS = ["post_qs", "post_body", "get_qs", "get_cookie", "get_uri", "get_header"]
PLACEMENT_LABEL = {
    "post_qs": "POST-QS", "post_body": "POST-Body",
    "get_qs": "GET-QS", "get_cookie": "GET-Cookie",
    "get_uri": "GET-URI", "get_header": "GET-Header",
}
POLICIES = ["1.1", "1.3", "1.3b"]

RULE_GROUPS = [
    ("Common",          "AWSManagedRulesCommonRuleSet",          "?", "Version_1.20"),
    ("SQLi",            "AWSManagedRulesSQLiRuleSet",            "?", "Version_2.0"),
    ("KnownBadInputs",  "AWSManagedRulesKnownBadInputsRuleSet",  "?", "Version_1.22"),
    ("Windows",         "AWSManagedRulesWindowsRuleSet",         "?", "Version_2.3"),
    ("Linux",           "AWSManagedRulesLinuxRuleSet",           "?", "Version_2.6"),
    ("Unix",            "AWSManagedRulesUnixRuleSet",            "?", "Version_3.2"),
    ("PHP",             "AWSManagedRulesPHPRuleSet",             "?", "Version_2.1"),
    ("AdminProtection", "AWSManagedRulesAdminProtectionRuleSet", "?", "Version_1.1"),
    ("WordPress",       "AWSManagedRulesWordPressRuleSet",       "?", "Version_1.3"),
    ("cyberwasp",       "baseline-1-3-cyberwasp-custom-rules",   "?", "present"),
    ("owasp-top10",     "baseline-1-3-owasp-top10-v3",           "?", "present"),
]

FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")
FILL_HEADER = PatternFill("solid", fgColor="305496")
FILL_CAVEAT = PatternFill("solid", fgColor="FFD966")
FONT_HEADER = Font(bold=True, color="FFFFFF")
FONT_BOLD   = Font(bold=True)


def parse_rule_short(rule_id):
    if not rule_id:
        return ""
    for short, full, *_ in RULE_GROUPS:
        if full in str(rule_id):
            return short
    return ""


def load_run(json_path):
    if not os.path.exists(json_path):
        return None
    d = json.loads(open(json_path).read(), strict=False)
    out = []
    for r in d["evidence"]:
        cf = r.get("cf_waf_log") or {}
        out.append({
            "test_id": str(r.get("requirement_id", "")),
            "request_id": r.get("request_id", ""),
            "verdict": r["verdict"],
            "rule_short": parse_rule_short(cf.get("terminating_rule_id")),
            "rule_id": cf.get("terminating_rule_id") or "",
            "cf_action": cf.get("action", ""),
            "cf_timestamp": cf.get("timestamp", ""),
        })
    return out


TEST_METADATA_CSV = os.path.join(HERE, "test_metadata.csv")


def load_test_metadata():
    """test_id -> {payload, attack_family} from test_metadata.csv (bundled with the repo).
    Unknown test_ids fall back to attack_family='Uncategorized'."""
    import csv
    out = {}
    if not os.path.exists(TEST_METADATA_CSV):
        print(f"WARN: {TEST_METADATA_CSV} not found; payload + attack_family columns will be blank")
        return out
    with open(TEST_METADATA_CSV, newline="") as fh:
        for r in csv.DictReader(fh):
            out[r["test_id"]] = {
                "payload": r.get("payload", ""),
                "attack_family": (r.get("attack_family") or "").strip().lower(),
            }
    return out


def write_raw_tabs(wb, all_data, lookup):
    raw_names = {}
    for (pol, p), records in all_data.items():
        sn = f"RAW_{pol}_{p.upper()}"[:31]
        raw_names[(pol, p)] = sn
        ws = wb.create_sheet(sn)
        hdrs = ["test_id", "request_id", "verdict", "rule_short", "rule_id",
                "payload", "cf_action", "cf_timestamp", "attack_family"]
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
        if records is None:
            continue
        for ri, rec in enumerate(records, start=2):
            lk = lookup.get(rec["test_id"], {})
            ws.cell(row=ri, column=1, value=rec["test_id"])
            ws.cell(row=ri, column=2, value=rec["request_id"])
            ws.cell(row=ri, column=3, value=rec["verdict"])
            ws.cell(row=ri, column=4, value=rec["rule_short"])
            ws.cell(row=ri, column=5, value=rec["rule_id"])
            ws.cell(row=ri, column=6, value=lk.get("payload", ""))
            ws.cell(row=ri, column=7, value=rec["cf_action"])
            ws.cell(row=ri, column=8, value=rec["cf_timestamp"])
            ws.cell(row=ri, column=9, value=lk.get("attack_family", "") or "Uncategorized")
        for col, w in [("A", 16), ("B", 45), ("C", 22), ("D", 18), ("E", 50),
                       ("F", 40), ("I", 14)]:
            ws.column_dimensions[col].width = w
    return raw_names


def write_definitions(ws, run_date):
    ws.cell(row=1, column=1, value="How to read this workbook").font = Font(bold=True, size=14)
    iso = f"{run_date[:4]}-{run_date[4:6]}-{run_date[6:]}"
    rows = [
        ("", ""),
        ("Data date", f"Generated {iso} from validate_baseline_*.json (1.1, 1.3, 1.3b across 6 placements: POST-QS, POST-Body, GET-QS, GET-Cookie, GET-URI, GET-Header)."),
        ("", ""),
        ("Tab order", ""),
        ("Summary", "Counts via COUNTIF formulas + CF_BLOCK share tables. Right-side block 'Where the leaks live' breaks down each (policy, placement) leak count by attack family (xss/sqli/cmdi/lfi/rfi/Uncategorized)."),
        ("Common, SQLi, ...", "Per-ruleset analytics. Each row is a test_id. Columns C-H = 1.1 vs 1.3. Columns I-N = 1.1 vs 1.3b."),
        ("RAW_<policy>_<placement>", "Raw data (18 tabs at the back) feeding the analytics."),
        ("", ""),
        ("Category values", ""),
        ("both", "Blocked by THIS ruleset in BOTH policies"),
        ("1.3 only", "Coverage gain - blocked by this ruleset in target policy but not in 1.1"),
        ("1.1 only (REGRESSION)", "RED - blocked in 1.1, NOT blocked at all in target policy"),
        ("(blank)", "Not relevant for this ruleset (caught by another rule, not tested, or both policies failed identically). Attribution shifts are intentionally blank in this view."),
    ]
    rows += [
        ("", ""),
        ("Verify any cell", ""),
        ("Applies to", "The 11 per-ruleset tabs. Summary is a roll-up; RAW_* tabs are the inputs Step 2 sends you to."),
        ("Step 1", "Pick the cell. Note test_id (col A) and placement from column header (C-H = 1.1 vs 1.3, I-N = 1.1 vs 1.3b)."),
        ("Step 2", "Open the two matching RAW tabs. Example: the POST-QS column under '1.1 vs 1.3b' -> RAW_1.1_POST_QS and RAW_1.3b_POST_QS."),
        ("Step 3", "Read request_id (col B) and rule (col D). Paste request_id into CW Logs Insights: fields @timestamp, action, terminatingRuleId | filter @message like /<request_id>/"),
    ]
    for r, (k, v) in enumerate(rows, start=2):
        c = ws.cell(row=r, column=1, value=k)
        if k in ("both", "1.3 only"):
            c.fill = FILL_GREEN; c.font = FONT_BOLD
        elif "REGRESSION" in k:
            c.fill = FILL_RED; c.font = FONT_BOLD
        elif k in ("Tab order", "Category values", "Verify any cell"):
            c.font = Font(bold=True, size=12)
        elif k in ("Data date", "Applies to"):
            c.font = Font(bold=True, italic=True); c.fill = FILL_CAVEAT
        ws.cell(row=r, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
        if v and len(str(v)) > 100:
            ws.row_dimensions[r].height = max(30, (len(str(v)) // 100) * 16 + 16)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110


def nested_if_formula(test_cell, raw_a, raw_b, rule_short):
    rule_a = f'IFERROR(VLOOKUP({test_cell},{raw_a}!$A:$J,4,FALSE),"")'
    rule_b = f'IFERROR(VLOOKUP({test_cell},{raw_b}!$A:$J,4,FALSE),"")'
    verdict_b = f'IFERROR(VLOOKUP({test_cell},{raw_b}!$A:$J,3,FALSE),"")'
    return (f'=IF(AND({rule_a}="{rule_short}",{rule_b}="{rule_short}"),"both",'
            f'IF(AND({rule_a}<>"{rule_short}",{rule_b}="{rule_short}"),"1.3 only",'
            f'IF(AND({rule_a}="{rule_short}",{rule_b}<>"{rule_short}",{verdict_b}<>"CF_BLOCK"),"1.1 only (REGRESSION)","")))')


def _ordered_families(all_tids, lookup):
    """Stable family list, frequency-sorted, Uncategorized last."""
    cnt = Counter()
    for tid in all_tids:
        fam = (lookup.get(tid, {}).get("attack_family") or "").strip().lower() or "Uncategorized"
        cnt[fam] += 1
    fams = sorted(cnt.items(), key=lambda x: -x[1])
    head = [k for k, _ in fams if k != "Uncategorized"]
    tail = ["Uncategorized"] if "Uncategorized" in cnt else []
    return head + tail


def write_leak_block(ws, start_row, start_col, all_tids, lookup):
    """C1-only block 'Where the leaks live - by attack family'. Live COUNTIFS formulas."""
    families = _ordered_families(all_tids, lookup)
    ws.cell(row=start_row, column=start_col,
            value="Where the leaks live - by attack family (counts of leaked tests)").font = Font(bold=True, size=12)
    hdr = ["Policy", "Placement"] + families + ["TOTAL leaked"]
    for c, h in enumerate(hdr, start_col):
        cell = ws.cell(row=start_row + 1, column=c, value=h)
        cell.font = FONT_HEADER; cell.fill = FILL_HEADER
    ri = start_row + 2
    for pol in POLICIES:
        for p in PLACEMENTS:
            raw = f"RAW_{pol}_{p.upper()}"
            ws.cell(row=ri, column=start_col, value=pol)
            ws.cell(row=ri, column=start_col + 1, value=PLACEMENT_LABEL[p])
            for fi, fam in enumerate(families, start=start_col + 2):
                ws.cell(row=ri, column=fi,
                        value=f'=COUNTIFS({raw}!C:C,"<>CF_BLOCK",{raw}!I:I,"{fam}")')
            first = get_column_letter(start_col + 2)
            last = get_column_letter(start_col + 1 + len(families))
            total_col = start_col + 2 + len(families)
            ws.cell(row=ri, column=total_col, value=f"=SUM({first}{ri}:{last}{ri})")
            ri += 1
    return ri  # next free row


def build(src_dir, out_dir, run_date):
    lookup = load_test_metadata()
    all_data = {}
    for pol in POLICIES:
        for p in PLACEMENTS:
            all_data[(pol, p)] = load_run(os.path.join(src_dir, f"validate_baseline_{pol}_{p}.json"))

    all_tids = set()
    for records in all_data.values():
        if records:
            for r in records:
                all_tids.add(r["test_id"])
    all_tids = sorted(
        all_tids,
        key=lambda t: int(re.search(r"\d+", t).group()) if re.search(r"\d+", t) else 99999,
    )

    wb = Workbook()
    df = wb.active; df.title = "Definitions"
    write_definitions(df, run_date)

    n = len(PLACEMENTS)
    b2_start = 3 + n              # first column of the "1.1 vs 1.3b" block
    total_cols = 2 + 2 * n        # Test ID + Payload + N + N

    rule_sheets = []
    for rule_short, rule_full, ver1, ver3 in RULE_GROUPS:
        ws = wb.create_sheet(rule_short[:31])
        rule_sheets.append(rule_short[:31])
        ws.cell(row=1, column=1, value=f"{rule_full} | 1.1: {ver1} -> 1.3: {ver3}").font = Font(bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws.cell(row=2, column=3, value="1.1 vs 1.3").font = FONT_BOLD
        ws.cell(row=2, column=b2_start, value="1.1 vs 1.3b").font = FONT_BOLD
        ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=2 + n)
        ws.merge_cells(start_row=2, start_column=b2_start, end_row=2, end_column=total_cols)
        col_hdrs = ["Test ID", "Payload"] + [PLACEMENT_LABEL[p] for p in PLACEMENTS] * 2
        for c, h in enumerate(col_hdrs, 1):
            cell = ws.cell(row=3, column=c, value=h); cell.font = FONT_HEADER; cell.fill = FILL_HEADER

        for ri, tid in enumerate(all_tids, start=4):
            ws.cell(row=ri, column=1, value=tid)
            ws.cell(row=ri, column=2, value=lookup.get(tid, {}).get("payload", ""))
            for ci, p in enumerate(PLACEMENTS, start=3):
                ws.cell(row=ri, column=ci, value=nested_if_formula(
                    f'"{tid}"', f"RAW_1.1_{p.upper()}", f"RAW_1.3_{p.upper()}", rule_short))
            for ci, p in enumerate(PLACEMENTS, start=b2_start):
                ws.cell(row=ri, column=ci, value=nested_if_formula(
                    f'"{tid}"', f"RAW_1.1_{p.upper()}", f"RAW_1.3b_{p.upper()}", rule_short))

        last_row = 3 + len(all_tids)
        rng = f"C4:{get_column_letter(total_cols)}{last_row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"both"'], fill=FILL_GREEN))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"1.3 only"'], fill=FILL_GREEN))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"1.1 only (REGRESSION)"'], fill=FILL_RED))

        widths = [16, 40] + [14] * (2 * n)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "C4"

    sm = wb.create_sheet("Summary")
    sm.cell(row=1, column=1, value="Summary - per-ruleset counts (formulas)").font = Font(bold=True, size=14)
    hdrs = ["Ruleset", "Comparison", "both", "1.3 only", "1.1 only (REGRESSION)"]
    for c, h in enumerate(hdrs, 1):
        cell = sm.cell(row=3, column=c, value=h); cell.font = FONT_HEADER; cell.fill = FILL_HEADER
    b1_first, b1_last = get_column_letter(3), get_column_letter(2 + n)
    b2_first, b2_last = get_column_letter(b2_start), get_column_letter(total_cols)
    rr = 4
    for rule_short, *_ in RULE_GROUPS:
        sn = rule_short[:31]
        last = 3 + len(all_tids)
        for cmp_label, col_range in [("1.1 vs 1.3", f"${b1_first}$4:${b1_last}"),
                                     ("1.1 vs 1.3b", f"${b2_first}$4:${b2_last}")]:
            sm.cell(row=rr, column=1, value=rule_short)
            sm.cell(row=rr, column=2, value=cmp_label)
            sm.cell(row=rr, column=3, value=f"=COUNTIF('{sn}'!{col_range}{last},\"both\")")
            sm.cell(row=rr, column=4, value=f"=COUNTIF('{sn}'!{col_range}{last},\"1.3 only\")")
            c5 = sm.cell(row=rr, column=5, value=f"=COUNTIF('{sn}'!{col_range}{last},\"1.1 only (REGRESSION)\")")
            c5.fill = FILL_RED
            rr += 1

    for i, w in enumerate([22, 14, 10, 12, 26], 1):
        sm.column_dimensions[get_column_letter(i)].width = w

    n_tests = len(all_tids)
    grand_total = n_tests * n
    rr += 2
    sm.cell(row=rr, column=1, value=f"CF_BLOCK share by managed ruleset (% of {grand_total} = {n_tests} tests x {n} placements)").font = Font(bold=True, size=12)
    rr += 1
    hdr2 = ["Ruleset", "1.1", "1.3", "1.3b"]
    for c, h in enumerate(hdr2, 1):
        cell = sm.cell(row=rr, column=c, value=h); cell.font = FONT_HEADER; cell.fill = FILL_HEADER
    rr += 1

    def _ruleset_share_formula(pol, rule_short):
        parts = []
        for p in PLACEMENTS:
            raw = f"RAW_{pol}_{p.upper()}"
            parts.append(f'COUNTIFS({raw}!D:D,"{rule_short}",{raw}!C:C,"CF_BLOCK")')
        return f"=({' + '.join(parts)})/{grand_total}"

    for rule_short, *_ in RULE_GROUPS:
        sm.cell(row=rr, column=1, value=rule_short)
        for ci, pol in enumerate(POLICIES, start=2):
            c = sm.cell(row=rr, column=ci, value=_ruleset_share_formula(pol, rule_short))
            c.number_format = "0.0%"
        rr += 1

    rr += 2
    sm.cell(row=rr, column=1, value=f"CF_BLOCK share by ruleset and placement (% of {n_tests} per placement)").font = Font(bold=True, size=12)
    rr += 1
    hdr3 = ["Ruleset", "Placement", "1.1", "1.3", "1.3b"]
    for c, h in enumerate(hdr3, 1):
        cell = sm.cell(row=rr, column=c, value=h); cell.font = FONT_HEADER; cell.fill = FILL_HEADER
    rr += 1

    def _placement_share_formula(pol, rule_short, placement):
        raw = f"RAW_{pol}_{placement.upper()}"
        return f'=COUNTIFS({raw}!D:D,"{rule_short}",{raw}!C:C,"CF_BLOCK")/{n_tests}'

    for rule_short, *_ in RULE_GROUPS:
        for plc in PLACEMENTS:
            sm.cell(row=rr, column=1, value=rule_short)
            sm.cell(row=rr, column=2, value=PLACEMENT_LABEL[plc])
            for ci, pol in enumerate(POLICIES, start=3):
                c = sm.cell(row=rr, column=ci, value=_placement_share_formula(pol, rule_short, plc))
                c.number_format = "0.0%"
            rr += 1

    # Leak-by-attack-family block, placed to the right of the Summary tables
    # (start at column H so the existing tables in A-E are untouched)
    write_leak_block(sm, start_row=3, start_col=8, all_tids=all_tids, lookup=lookup)
    for c in range(8, 18):
        sm.column_dimensions[get_column_letter(c)].width = 14

    raw_names = write_raw_tabs(wb, all_data, lookup)
    order = (["Definitions", "Summary"] + rule_sheets
             + [raw_names[(pol, p)] for pol in POLICIES for p in PLACEMENTS])
    wb._sheets = [wb[n] for n in order if n in wb.sheetnames]
    wb.calculation = CalcProperties(calcId=124519, fullCalcOnLoad=True)

    out = os.path.join(out_dir, f"comparison_1.1_vs_1.3_vs_1.3b_FORMULAS_FIXED_{run_date}.xlsx")
    wb.save(out)
    return out


def main():
    p = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    p.add_argument("--date", required=True, help="run date YYYYMMDD (used in output filename)")
    p.add_argument("--src-dir", default=None,
                   help="dir containing validate_baseline_*.json (default: runs/<date>_run/)")
    p.add_argument("--out-dir", default=os.path.join(HERE, "runs"),
                   help="output dir for comparison xlsx (default: runs/)")
    args = p.parse_args()

    src_dir = args.src_dir or os.path.join(HERE, "runs", f"{args.date}_run")
    if not os.path.isdir(src_dir):
        print(f"ERROR: source dir {src_dir} not found")
        sys.exit(1)
    os.makedirs(args.out_dir, exist_ok=True)

    print(f"Source: {src_dir}")
    print(f"Output: {args.out_dir}")
    print()

    out = build(src_dir, args.out_dir, args.date)
    print(f"  Wrote {out}")


if __name__ == "__main__":
    main()
